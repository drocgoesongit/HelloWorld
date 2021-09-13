import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.open('transactions.xlsx')
sheet = wb['Sheet1']

for rows in range(2, sheet.max_row + 1):
    cell = sheet.cell(rows, 3)
    corrected_value = cell.value * 0.9
    new_cell = sheet.cell(rows, 4)
    new_cell.value = corrected_value

cell = sheet.cell(1, 4)
cell.value = 'Corrected Price'

chart1 = BarChart()
data = Reference(sheet, min_row=2, max_row=sheet.max_row,min_col=4, max_col=4)
chart1.add_data(data)
sheet.add_chart(chart1, 'a7')

wb.save('transactions2.xlsx')
